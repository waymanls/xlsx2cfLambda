#!/usr/bin/env python
import os
import sys
import json
from jinja2.environment import Template
import yaml
import boto3
import argparse
from datetime import date
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader

# Read in EC2 Resource block
with open("EC2.template.jinja2", 'r') as f:
	EC2Block = f.read()
# Read in RDS Resource block
with open("RDS.template.jinja2", 'r') as f1:
	RDSBlock = f1.read()

def getDatafromSheet(row, ticketNumber):
	"""
	Query spreadsheet for either hostname or Instance name and pull relevant fields
	"""
	#workbook = load_workbook(filename = spreadsheet)
	resourceDict = {}
	#EC2worksheet = workbook.active
	#session = boto3.session.Session()

	#for row in EC2worksheet.iter_rows(min_row=2,values_only=True):
	resourceDict['ServerType'] = row[0]
	resourceDict['ResourceType'] = row[1]
	resourceDict['ArchitectureID'] = row[2]
	resourceDict['AvailabilityZone'] = row[3]
	resourceDict['AmiId'] = row[4]
	resourceDict['OS'] = row[5]
	resourceDict['Version'] = row[6]
	resourceDict['InstanceType'] = row[7]
	resourceDict['RootVolSize'] = row[8]
	resourceDict['CloudEnvironment'] = row[9]
	resourceDict['Subnet'] = row[10]
	resourceDict['Environment'] = row[11]
	resourceDict['InstanceName'] = row[12]
	resourceDict['Hostname'] = row[13]
	resourceDict['WebAdaptorName'] = row[14]
	resourceDict['AdditionalVolSize'] = row[15]
	resourceDict['MissionOwner'] = row[16]
	resourceDict['Office'] = row[17]
	resourceDict['Product'] = row[18]
	resourceDict['Startup'] = row[19]
	resourceDict['Shutdown'] = row[20]
	resourceDict['Schedule'] = row[21]
	resourceDict['SecurityGroup'] = row[22]
#		if name == row[9] or name == row[10]:
#			resourceDict['Domain'] = row[0]
#			resourceDict['Environment'] = row[1]
#			resourceDict['MissionOwner'] = row[2]
#			resourceDict['Office'] = row[3]
#			resourceDict['Product'] = row[4]
#			resourceDict['ArchitectureId'] = row[6]
#			resourceDict['Name'] = row[9]
#			resourceDict['Hostname'] = row[10]
#			resourceDict['ServerType'] = row[11]
#			resourceDict['VolSize'] = row[12]
#			resourceDict['InstanceType'] = row[16]
#			resourceDict['SubnetId'] = row[18]
#			resourceDict['Schedule'] = row[22]
#			resourceDict['Startup'] = row[23]
#			resourceDict['Shutdown'] = row[24]
	if not resourceDict:
		print("=================================================================================")
		print("                                     ERROR                                       ")
		print("=================================================================================")
		print(name + " not found in spreadsheet. Please verify that name is correct")
		print("")
		print("")
		print("Exiting...")
		print("=================================================================================")
		sys.exit()
#	else:
#		# Translate SubnetName to Id since we can evaluate formulas that define subnets in spreadsheet
#		resourceDict['JiraTicket'] = ticketNumber
#		ec2 = boto3.resource('ec2')
#		filters = [{'Name':'tag:Name','Values':[resourceDict['SubnetId']]}]
#		subnet = list(ec2.subnets.filter(Filters=filters))
#		resourceDict['SubnetId'] = str(subnet[0]).split('\'')[1]

	return resourceDict

#def genTemplate(resourceDict):
def genTemplate(spreadsheet):
	"""
	Generate Cloudformatin template using jinja2 using values from Spreadsheet
	"""
	env = Environment(loader = FileSystemLoader('.'), trim_blocks=True, lstrip_blocks=True)
	Servertemplate = env.get_template('Server.template.jinja2')
	EC2template = env.get_template('EC2.template.jinja2')
	RDStemplate = env.get_template('RDS.template.jinja2')
	Template = ""
	workbook = load_workbook(filename = spreadsheet)
	EC2worksheet = workbook.active
	#print(EC2worksheet['B2'].value)
	# Print each row, and generate a template based on the value in column
	# if resource is defined in spreadsheet include it in template, otherwise, don't
	for row in EC2worksheet.iter_rows(min_row=2,values_only=True):
            templateValues = getDatafromSheet(row, workbook.active.title)
            if row[1] == 'EC2':
				# Get values from spreadsheet
				# Render EC2 Resource Block
                EC2BlockTemplate = EC2template.render(templateValues)
                # Append to EC2 resource block to Resources template
                Template+=EC2BlockTemplate
            if row[1] == 'RDS':
				# Get values from spreadsheet
				# Render EC2 Resource Block
                RDSBlockTemplate = RDStemplate.render(templateValues)
                # Append to EC2 resource block to Resources template
                Template+=RDSBlockTemplate
	with open('Final.template.jinja2', 'w') as f3:
		f3.write("{% block content %}\n")
		f3.write("{% endblock content %}\n")
		f3.write(Template)
		f3.write("\n")
	cfTemplate = Servertemplate.render(templateValues)
	print(cfTemplate)

	#return cfTemplate

def buildStack(cfTemplate,ticketNumber,resourceName):
	"""
	Create CloudFormation stack using template generated in genTemplate()
	"""
	stackName = resourceName + "-" + ticketNumber + "-" + str(date.today())
	client = boto3.client('cloudformation')
	parameterFile = open("GovCloud-Automation-Test-Servers.json","r")
	params = json.load(parameterFile)
	response = client.create_stack(StackName=stackName,TemplateBody=cfTemplate,DisableRollback=True,Parameters=params)
	print("==============================================================================")
	print("                                  SUCCESS                                     ")
	print("==============================================================================")
	print("Stack has been created, please track progress in CloudFormation Web Console")
	print("==============================================================================")
	print(response['StackId'])
	print("==============================================================================")
	parameterFile.close()

def fileCheck():
	"""
	Check to see if jinja template and parameter files are in place
	"""
	if not os.path.exists("GovCloud-Automation-Test-Servers.json") or not os.path.exists("Server.template.jinja2"):
		sys.exit()
def main():
	parser = argparse.ArgumentParser(description="Tool to create server from spreadsheet via CloudFormation")
	parser.add_argument('-S', '--spreadsheet', help='Path to spreadsheet') 	
	args = parser.parse_args()
	#cfTemplate = genTemplate(args.spreadsheet)
	cfTemplate = genTemplate('FG-47856.xlsx')

if __name__ == '__main__':
	main()
