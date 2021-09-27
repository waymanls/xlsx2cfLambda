#!/usr/bin/env python
import os
import sys
import json
import yaml
import boto3
import argparse
import botocore.exceptions
import urllib.parse
from jinja2.environment import Template
from datetime import date
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader

s3 = boto3.client('s3')
s33 = boto3.resource('s3')

# Read in EC2 Resource block
with open("EC2.template.jinja2", 'r') as f:
	EC2Block = f.read()
# Read in RDS Resource block
with open("RDS.template.jinja2", 'r') as f1:
	RDSBlock = f1.read()

def getDatafromSheet(row, ticketNumber):
	"""
	Query spreadsheet for either hostname or Instance name and pull relevant fields
	"""
	resourceDict = {}
	resourceDict['ServerType'] = row[0]
	resourceDict['ResourceType'] = row[1]
	resourceDict['ArchitectureID'] = row[2]
	resourceDict['AvailabilityZone'] = row[3]
	resourceDict['AmiId'] = row[4]
	resourceDict['OS'] = row[5]
	resourceDict['Version'] = row[6]
	resourceDict['InstanceType'] = row[7]
	resourceDict['StorageType'] = row[8]
	resourceDict['Backups'] = int(row[9])
	resourceDict['RootVolSize'] = int(row[10])
	resourceDict['CloudEnvironment'] = row[11]
	resourceDict['SubnetId'] = row[12]
	resourceDict['Environment'] = row[13]
	resourceDict['InstanceName'] = row[14]
	resourceDict['Hostname'] = row[15]
	resourceDict['WebAdaptorName'] = row[16]
	resourceDict['AdditionalVolSize'] = row[17]
	resourceDict['MissionOwner'] = row[18]
	resourceDict['Office'] = row[19]
	resourceDict['Product'] = row[20]
	resourceDict['Startup'] = row[21]
	resourceDict['Shutdown'] = row[22]
	resourceDict['Schedule'] = row[23]
	resourceDict['SecurityGroup'] = row[24]
	resourceDict['WorkloadType'] = row[25]
	resourceDict['JiraTicket'] = ticketNumber

	return resourceDict

def genTemplate(spreadsheet):
	"""
	Generate Cloudformatin template using jinja2 using values from Spreadsheet
	"""
	env = Environment(loader = FileSystemLoader(['.','/tmp']), trim_blocks=True, lstrip_blocks=True)
	Servertemplate = env.get_template('Server.template.jinja2')
	EC2template = env.get_template('EC2.template.jinja2')
	RDStemplate = env.get_template('RDS.template.jinja2')
	Template = ""
	workbook = load_workbook(filename = spreadsheet)
	EC2worksheet = workbook.active
	
	for row in EC2worksheet.iter_rows(min_row=2,values_only=True):
            templateValues = getDatafromSheet(row, workbook.active.title)
            if row[1] == 'EC2':
                EC2BlockTemplate = EC2template.render(templateValues)
                # Append to EC2 resource block to Resources template
                Template+=EC2BlockTemplate
            if row[1] == 'RDS':
                RDSBlockTemplate = RDStemplate.render(templateValues)
                # Append to EC2 resource block to Resources template
                Template+=RDSBlockTemplate
	with open('/tmp/Final.template.jinja2', 'w') as f3:
		f3.write("{% block content %}\n")
		f3.write("{% endblock content %}\n")
		f3.write(Template)
		f3.write("\n")
	cfTemplate = Servertemplate.render(templateValues)

	return cfTemplate, workbook.active.title

def buildStack(cfTemplate,jiraTicket):
	"""
	Create CloudFormation stack using template generated in genTemplate()
	"""

	client = boto3.client('cloudformation')
	stack_name = jiraTicket + '-' + date.isoformat(date.today())
	try:
		response = client.validate_template(TemplateBody=cfTemplate)
	except botocore.exceptions.ClientError as err:
		print(err)
	try:
		response = client.create_stack(StackName=stack_name,TemplateBody=cfTemplate,DisableRollback=True)
	except botocore.exceptions.ClientError as err:
		print(err)

	return response

def lambda_handler(event, context):

    # Get the object from the event and show its content type
    bucket = event['Records'][0]['s3']['bucket']['name']
    key = urllib.parse.unquote_plus(event['Records'][0]['s3']['object']['key'], encoding='utf-8')
    try:
        response = s3.get_object(Bucket=bucket, Key=key)
        s33.meta.client.download_file(bucket,key,'/tmp/sample.xlsx')
        cfTemplate, jiraTicket = genTemplate('/tmp/sample.xlsx')
        buildStack(cfTemplate,jiraTicket)
        return cfTemplate
    except Exception as e:
        print(e)
        print('Error getting object {} from bucket {}. Make sure they exist and your bucket is in the same region as this function.'.format(key, bucket))
        raise e

